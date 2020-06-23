import openpyxl
import os
import excel_utils
import logging
from log_utils import logd, logi, logw, loge, set_log_level


set_log_level(logging.ERROR)

# filename = 'test/random.xlsx'
summary = 'E:\gp\\6.xlsx'
date = '6.23'
file = 'E:\gp\origin\\' + date + '.xlsx'

if not os.path.exists(summary):
    excel_utils.create_excel(summary)

def process_excel(finename, sheetname, filetemp):
    wb = openpyxl.load_workbook(finename)
    # sheet = wb.active
    sheet = wb.create_sheet(sheetname)
    loge('start file %s------------------------', filetemp)
    wb_temp = openpyxl.load_workbook(filetemp)
    sheet_temp = wb_temp.active

    sheet_temp = excel_utils.find_and_delete_key(sheet_temp, 'SH688', 1)
    sheet_temp = excel_utils.find_and_delete_key(sheet_temp, 'ST', 1)
    sheet_temp = excel_utils.delete_range(sheet_temp, 9)
    # for row in range(2, sheet_temp.max_row + 1):
    #     logi('%s, %s', sheet_temp.cell(row, 1).value, sheet_temp.cell(row, 2).value)
    for row in range(1, sheet_temp.max_row+1):
        for cell in sheet_temp[row]:
            sheet.cell(row, cell.column).value = cell.value
# wb.create_sheet('new', 2)
    wb.save(summary)

process_excel(summary, date, file)
