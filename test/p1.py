import openpyxl

filename = 'D:\moba_cts_test.xlsx'
new_file = 'D:\moba_cts_test_1.xls'

wb = openpyxl.load_workbook(filename)

# print(wb.sheetnames)

# for sheet in wb:
#     print(sheet.title)
#
# newSheet = wb.create_sheet('Sheet3')
#
# print(wb.sheetnames)
#
# sheet3 = wb.get_sheet_by_name('Sheet3')
# sheet3 = wb['Sheet3']

ws = wb.active

# print(ws)
# print(ws['A1'])
# print(ws['A1'].value)
# c = ws['B1']
# print('Row {}, Column {} is {}' .format(c.row, c.column, c.value))
# print('Cell {}, is {}' .format(c.coordinate, c.value))
# print(ws.cell(row=1, column=2))
# print(ws.cell(row=1, column=2).value)

# colC = ws['C']
# for col in colC:
#     print(col.value)
#
# row6 = ws[6]

col_range = ws['B:C']
row_range = ws[2:6]
# for col in col_range:
#     for cell in col:
#         print(cell.value)
#
# for row in row_range:
#     for cell in row:
#         print(cell.value)

# for row in ws.iter_rows(min_row=1, max_row=2, max_col=2):
#     for cell in row:
#         print(cell.value)

# print(tuple(ws.rows))
import random

cell_range = ws['A1:C20']
for rowOfCellObject in cell_range:
    for cellObj in rowOfCellObject:
        cellObj.value = random.randint(0, 999999)
        print(cellObj.coordinate, cellObj.value)
wb.save('random.xlsx')
# print('{}, {}' .format(ws.max_row, ws.max_column))

# from openpyxl.utils import get_column_letter, column_index_from_string
# print(get_column_letter(2), get_column_letter(25))


