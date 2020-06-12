import openpyxl

'''
{

'''
filename = filename = 'random.xlsx'

wb = openpyxl.load_workbook(filename)
sheet = wb.active

countryDate = []
# print(sheet.max_row)
for row in range(2, sheet.max_row+1):
    for cell in sheet[str(row)]:
        data = cell.value
        if '32223' in str(data):
            sheet.delete_rows(row)
for row in range(2, sheet.max_row+1):
    print(sheet.cell(row, 1).value)


