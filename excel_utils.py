import openpyxl
from log_utils import logd, logi, logw, loge, set_log_level


def find_and_delete_key(workSheet, key, col=0):
    # print(sheet.max_row)
    flag = 0
    for row in range(workSheet.max_row, 1, -1):
        #不指定是哪一列，轮询所有列
        if col == 0:
            for cell in workSheet[row]:
                data = cell.value
                if key in str(data):
                    flag = 1
                    break
        #只判断某一列
        else:
            data = workSheet.cell(row, col).value
            if key in str(data):
                flag = 1
        if flag == 1:
            logi('delete %s, %s', workSheet.cell(row, 1).value, workSheet.cell(row, 2).value)
            workSheet.delete_rows(row)
            flag = 0

    # workSheet.max_row = workSheet.max_row - i
    return workSheet

def delete_range(workSheet, end, start=-99):
    flag = 0
    loge('max row %d', workSheet.max_row)
    for row in range(workSheet.max_row, 1, -1):
        data = workSheet.cell(row, 3).value
        if data == '--' or int(data) > start and data < end:
            flag = 1
        if flag == 1:
            logi('delete %s, %s, %s', workSheet.cell(row, 1).value, workSheet.cell(row, 2).value, workSheet.cell(row, 3).value)
            workSheet.delete_rows(row)
            flag = 0
    return workSheet

def create_excel(filename):
    wb = openpyxl.Workbook()
    wb.save(filename)
